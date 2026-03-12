from __future__ import annotations

import hashlib
import posixpath
import zipfile
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from fastapi import HTTPException


ROOT_DIR = Path(__file__).resolve().parents[2]
PRODUCT_MEDIA_DIR = ROOT_DIR / "media" / "products"
WORKBOOK_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

NAME_HEADER_PREFIXES = ("\u043d\u0430\u0437\u0432", "name", "\u0442\u043e\u0432\u0430\u0440")
PRICE_HEADER_PREFIXES = ("\u0446\u0435\u043d\u0430", "price", "\u0441\u0442\u043e\u0438\u043c")
FORMAT_HEADER_PREFIXES = ("\u0444\u043e\u0440\u043c\u0430\u0442", "size", "dimension")
CATEGORY_HEADER_PREFIXES = ("\u043a\u0430\u0442\u0435\u0433", "category")
THICKNESS_HEADER_PREFIXES = ("\u0442\u043e\u043b\u0449", "thickness")
IMAGE_HEADER_KEYWORDS = (
    "\u0444\u043e\u0442\u043e",
    "\u0438\u0437\u043e\u0431\u0440\u0430\u0436",
    "\u043a\u0430\u0440\u0442\u0438\u043d",
    "image",
    "photo",
    "img",
)


@dataclass(slots=True)
class EmbeddedImageAsset:
    data: bytes
    extension: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    order: int


def _normalized_header(value) -> str:
    if not isinstance(value, str):
        return ""

    return value.strip().casefold()


def _normalize_zip_target(base_path: str, target: str | None) -> str | None:
    if not target:
        return None

    normalized_target = target.replace("\\", "/")
    if normalized_target.startswith("/"):
        return normalized_target.lstrip("/")

    return posixpath.normpath(posixpath.join(posixpath.dirname(base_path), normalized_target))


def _active_sheet_path(archive: zipfile.ZipFile) -> str | None:
    workbook_path = "xl/workbook.xml"
    workbook_rels_path = "xl/_rels/workbook.xml.rels"

    if workbook_path not in archive.namelist() or workbook_rels_path not in archive.namelist():
        return None

    workbook_root = ET.fromstring(archive.read(workbook_path))
    workbook_view = workbook_root.find(f".//{{{WORKBOOK_NS}}}workbookView")

    try:
        active_index = int(workbook_view.attrib.get("activeTab", "0")) if workbook_view is not None else 0
    except ValueError:
        active_index = 0

    sheets_node = workbook_root.find(f".//{{{WORKBOOK_NS}}}sheets")
    if sheets_node is None:
        return None

    sheets = sheets_node.findall(f"{{{WORKBOOK_NS}}}sheet")
    if not sheets:
        return None

    active_sheet = sheets[active_index] if active_index < len(sheets) else sheets[0]
    relation_id = active_sheet.attrib.get(f"{{{DOC_REL_NS}}}id")
    if not relation_id:
        return None

    workbook_rels_root = ET.fromstring(archive.read(workbook_rels_path))
    for relation in workbook_rels_root:
        if relation.attrib.get("Id") == relation_id:
            return _normalize_zip_target(workbook_path, relation.attrib.get("Target"))

    return None


def _extract_embedded_images(file_bytes: bytes) -> list[EmbeddedImageAsset]:
    images: list[EmbeddedImageAsset] = []

    with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
        sheet_path = _active_sheet_path(archive)
        if not sheet_path:
            return images

        sheet_rels_path = posixpath.join(
            posixpath.dirname(sheet_path),
            "_rels",
            f"{posixpath.basename(sheet_path)}.rels",
        )
        if sheet_rels_path not in archive.namelist():
            return images

        sheet_rels_root = ET.fromstring(archive.read(sheet_rels_path))
        drawing_path = None
        for relation in sheet_rels_root:
            if relation.attrib.get("Type", "").endswith("/drawing"):
                drawing_path = _normalize_zip_target(sheet_path, relation.attrib.get("Target"))
                break

        if not drawing_path or drawing_path not in archive.namelist():
            return images

        drawing_rels_path = posixpath.join(
            posixpath.dirname(drawing_path),
            "_rels",
            f"{posixpath.basename(drawing_path)}.rels",
        )
        if drawing_rels_path not in archive.namelist():
            return images

        drawing_relations: dict[str, str] = {}
        drawing_rels_root = ET.fromstring(archive.read(drawing_rels_path))
        for relation in drawing_rels_root:
            relation_id = relation.attrib.get("Id")
            target_path = _normalize_zip_target(drawing_path, relation.attrib.get("Target"))
            if relation_id and target_path:
                drawing_relations[relation_id] = target_path

        namespaces = {
            "xdr": DRAWING_NS,
            "a": DRAWING_MAIN_NS,
            "r": DOC_REL_NS,
        }
        drawing_root = ET.fromstring(archive.read(drawing_path))
        anchors = drawing_root.findall("xdr:oneCellAnchor", namespaces)
        anchors.extend(drawing_root.findall("xdr:twoCellAnchor", namespaces))

        for order, anchor in enumerate(anchors):
            from_node = anchor.find("xdr:from", namespaces)
            to_node = anchor.find("xdr:to", namespaces)
            row_node = from_node.find("xdr:row", namespaces) if from_node is not None else None
            col_node = from_node.find("xdr:col", namespaces) if from_node is not None else None
            to_row_node = to_node.find("xdr:row", namespaces) if to_node is not None else None
            to_col_node = to_node.find("xdr:col", namespaces) if to_node is not None else None
            blip_node = anchor.find(".//a:blip", namespaces)

            if row_node is None or col_node is None or blip_node is None:
                continue

            relation_id = blip_node.attrib.get(f"{{{DOC_REL_NS}}}embed")
            media_path = drawing_relations.get(relation_id or "")
            if not media_path or media_path not in archive.namelist():
                continue

            try:
                start_row = int((row_node.text or "0").strip()) + 1
                start_col = int((col_node.text or "0").strip()) + 1
            except ValueError:
                continue

            try:
                end_row = int((to_row_node.text or "0").strip()) + 1 if to_row_node is not None else start_row
            except ValueError:
                end_row = start_row

            try:
                end_col = int((to_col_node.text or "0").strip()) + 1 if to_col_node is not None else start_col
            except ValueError:
                end_col = start_col

            extension = Path(media_path).suffix.lower().lstrip(".") or "png"
            if len(extension) > 5:
                extension = "png"

            images.append(
                EmbeddedImageAsset(
                    data=archive.read(media_path),
                    extension=extension,
                    start_row=start_row,
                    end_row=max(start_row, end_row),
                    start_col=start_col,
                    end_col=max(start_col, end_col),
                    order=order,
                )
            )

    return images


def _save_embedded_product_image(row_number: int, asset: EmbeddedImageAsset) -> str:
    PRODUCT_MEDIA_DIR.mkdir(parents=True, exist_ok=True)

    content_hash = hashlib.sha256(asset.data).hexdigest()[:16]
    filename = f"product-{row_number}-{content_hash}.{asset.extension}"
    target_path = PRODUCT_MEDIA_DIR / filename

    if not target_path.exists():
        target_path.write_bytes(asset.data)

    return f"/media/products/{filename}"


def _is_image_header(value: str) -> bool:
    return any(keyword == "img" and value == "img" or keyword != "img" and keyword in value for keyword in IMAGE_HEADER_KEYWORDS)


def _column_distance(asset: EmbeddedImageAsset, column_number: int | None) -> int:
    if column_number is None:
        return 0

    if asset.start_col <= column_number <= asset.end_col:
        return 0

    if column_number < asset.start_col:
        return asset.start_col - column_number

    return column_number - asset.end_col


def _row_assignment_cost(
    row_number: int,
    asset: EmbeddedImageAsset,
    image_column_number: int | None,
) -> int:
    center_row = (asset.start_row + asset.end_row) / 2
    distance_to_center = abs(row_number - center_row)

    if asset.start_row <= row_number <= asset.end_row:
        row_penalty = int(distance_to_center * 4)
    elif (
        abs(row_number - asset.start_row) <= 1
        or abs(row_number - asset.end_row) <= 1
        or distance_to_center <= 1
    ):
        row_penalty = 8 + int(distance_to_center * 4)
    else:
        row_penalty = 20 + int(distance_to_center * 8)

    return row_penalty + _column_distance(asset, image_column_number) * 15


def _assign_embedded_images_to_rows(
    product_row_numbers: list[int],
    image_column_number: int | None,
    images: list[EmbeddedImageAsset],
) -> dict[int, EmbeddedImageAsset]:
    if not product_row_numbers or not images:
        return {}

    row_numbers = sorted(product_row_numbers)
    preferred_images = sorted(images, key=lambda asset: (_column_distance(asset, image_column_number), asset.order))

    if image_column_number is not None:
        column_matched_images = [asset for asset in preferred_images if _column_distance(asset, image_column_number) <= 1]
        if column_matched_images:
            preferred_images = column_matched_images

    row_count = len(row_numbers)
    image_count = len(preferred_images)
    large_cost = 10**9
    skip_image_penalty = 25

    costs = [[large_cost] * (image_count + 1) for _ in range(row_count + 1)]
    parents: list[list[tuple[int, int, str] | None]] = [
        [None] * (image_count + 1) for _ in range(row_count + 1)
    ]
    costs[0][0] = 0

    for row_index in range(row_count + 1):
        for image_index in range(image_count + 1):
            current_cost = costs[row_index][image_index]
            if current_cost == large_cost:
                continue

            if row_index < row_count and current_cost < costs[row_index + 1][image_index]:
                costs[row_index + 1][image_index] = current_cost
                parents[row_index + 1][image_index] = (row_index, image_index, "skip_row")

            if image_index < image_count:
                skipped_image_cost = current_cost + skip_image_penalty
                if skipped_image_cost < costs[row_index][image_index + 1]:
                    costs[row_index][image_index + 1] = skipped_image_cost
                    parents[row_index][image_index + 1] = (row_index, image_index, "skip_image")

            if row_index < row_count and image_index < image_count:
                assigned_cost = current_cost + _row_assignment_cost(
                    row_numbers[row_index],
                    preferred_images[image_index],
                    image_column_number,
                )
                if assigned_cost < costs[row_index + 1][image_index + 1]:
                    costs[row_index + 1][image_index + 1] = assigned_cost
                    parents[row_index + 1][image_index + 1] = (row_index, image_index, "assign")

    assigned_rows: dict[int, EmbeddedImageAsset] = {}
    row_index = row_count
    image_index = image_count

    while row_index > 0 or image_index > 0:
        parent = parents[row_index][image_index]
        if parent is None:
            break

        previous_row_index, previous_image_index, action = parent
        if action == "assign":
            assigned_rows[row_numbers[previous_row_index]] = preferred_images[previous_image_index]

        row_index = previous_row_index
        image_index = previous_image_index

    return assigned_rows


def parse_products_by_names(file_bytes: bytes) -> list[dict]:
    """Read products from an xlsx file for bulk upsert."""
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active

    name_idx = -1
    price_idx = -1
    format_idx = -1
    category_idx = -1
    thickness_idx = -1
    image_idx = -1

    for idx, cell in enumerate(worksheet[1]):
        value = _normalized_header(cell.value)

        if value.startswith(NAME_HEADER_PREFIXES):
            name_idx = idx
        elif value.startswith(PRICE_HEADER_PREFIXES):
            price_idx = idx
        elif value.startswith(FORMAT_HEADER_PREFIXES):
            format_idx = idx
        elif value.startswith(CATEGORY_HEADER_PREFIXES):
            category_idx = idx
        elif value.startswith(THICKNESS_HEADER_PREFIXES):
            thickness_idx = idx
        elif _is_image_header(value):
            image_idx = idx

    if name_idx == -1 or price_idx == -1 or category_idx == -1:
        raise HTTPException(
            status_code=400,
            detail="\u041d\u0435\u0442 \u043e\u0434\u043d\u043e\u0439 \u0438\u0437 \u043d\u0435\u043e\u0431\u0445\u043e\u0434\u0438\u043c\u044b\u0445 \u043a\u043e\u043b\u043e\u043d\u043e\u043a",
        )

    product_row_numbers: list[int] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        raw_name = row[name_idx].value
        if raw_name is None:
            continue

        if isinstance(raw_name, str) and raw_name.strip() == "":
            continue

        product_row_numbers.append(row_number)

    embedded_images_by_row = _assign_embedded_images_to_rows(
        product_row_numbers=product_row_numbers,
        image_column_number=image_idx + 1 if image_idx != -1 else None,
        images=_extract_embedded_images(file_bytes),
    )

    products: list[dict] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        raw_name = row[name_idx].value
        if raw_name is None:
            continue

        if isinstance(raw_name, str) and raw_name.strip() == "":
            continue

        thickness = row[thickness_idx].value if thickness_idx != -1 else None
        price = Decimal(str(row[price_idx].value))
        raw_format = row[format_idx].value if format_idx != -1 else None
        category_name = row[category_idx].value
        category_name = category_name.strip() if isinstance(category_name, str) else category_name
        format_value = raw_format.strip() if isinstance(raw_format, str) else ""

        name = raw_name if thickness is None else f"{raw_name} {thickness} \u043c\u043c"

        image_url = None
        if image_idx != -1:
            image_cell = row[image_idx]
            if image_cell.hyperlink is not None and image_cell.hyperlink.target:
                image_url = image_cell.hyperlink.target.strip() or None
            elif isinstance(image_cell.value, str):
                image_url = image_cell.value.strip() or None

        if not image_url:
            embedded_image = embedded_images_by_row.get(row_number)
            if embedded_image is not None:
                image_url = _save_embedded_product_image(row_number, embedded_image)

        if "*" in format_value:
            width, length = map(str.strip, format_value.split("*", 1))
            max_width = int(width)
            max_length = int(length)
        else:
            max_width = None
            max_length = None

        products.append(
            {
                "name": name,
                "image_url": image_url,
                "thickness_mm": thickness,
                "price_per_m2": price,
                "max_width": max_width,
                "max_length": max_length,
                "category_name": category_name,
            }
        )

    return products


def parse_categories_of_products(file_bytes: bytes):
    """Read unique product categories from an xlsx file."""
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active

    categories_col_id = -1
    headers = [_normalized_header(cell.value) for cell in worksheet[1]]

    for index, header in enumerate(headers, start=1):
        if header.startswith(CATEGORY_HEADER_PREFIXES):
            categories_col_id = index
            break

    if categories_col_id == -1:
        raise HTTPException(
            status_code=400,
            detail="\u041d\u0435\u0442 \u043a\u043e\u043b\u043e\u043d\u043e\u043a \u0441 \u043a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u044f\u043c\u0438",
        )

    seen = set()
    categories = []
    for row in worksheet.iter_rows(
        min_row=2,
        min_col=categories_col_id,
        max_col=categories_col_id,
        values_only=True,
    ):
        if row[0] is None:
            continue

        category_name = str(row[0]).strip()
        if category_name == "" or category_name.casefold() in seen:
            continue

        categories.append({"category_name": category_name})
        seen.add(category_name.casefold())

    if not categories:
        raise HTTPException(
            status_code=400,
            detail="\u041d\u0435\u0442 \u043a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u0439 \u0442\u043e\u0432\u0430\u0440\u0430",
        )

    return categories
